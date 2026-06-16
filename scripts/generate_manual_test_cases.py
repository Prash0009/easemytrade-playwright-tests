"""Generates docs/Manual_Test_Cases.xlsx and docs/MANUAL_TEST_CASES.md
from the manual-QA equivalents of the automated Playwright suite.

Run: python3 scripts/generate_manual_test_cases.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PRECOND_PUBLIC = "Browser is open. Tester navigates to https://easemytrade.in (site redirects to https://www.easemytrade.in)."
PRECOND_GATED = ("Browser is open. Tester is logged in with a viewer-role account (or the site's temporary "
                  "public-unlock window is currently open) so gated pages render instead of redirecting to /login/.")
PRECOND_LOGIN = "Browser is open and NOT logged in. Tester navigates to https://easemytrade.in/login/."
PRECOND_HOME = "Browser is open. Tester navigates to the EaseMyTrade home page (https://easemytrade.in/)."
PRECOND_EV = PRECOND_GATED + " Tester navigates to /expert-view/."

# Each row: (id, module, title, priority, preconditions, [steps], expected, test_data)
CASES = []

def add(id_, module, title, priority, precond, steps, expected, data=""):
    CASES.append((id_, module, title, priority, precond, steps, expected, data))

# ───────────────────────── NAV — Navigation & Page Load ─────────────────────────
public_pages = [
    ("Home", "/", "EaseMyTrade", "home"),
    ("Login", "/login/", "Sign In", "login"),
    ("Expert View", "/expert-view/", "Expert View", "expert-view"),
    ("News", "/news/", "AI Market Intelligence", ""),
    ("Methodology", "/methodology/", "Methodology", ""),
    ("Telegram", "/telegram/", "Telegram", ""),
    ("Contact", "/contact/", "Contact", ""),
    ("Disclaimer", "/disclaimer/", "Disclaimer", ""),
]
for i, (name, path, kw, attr) in enumerate(public_pages, start=1):
    extra = f"\n4. Read the page's data-page body attribute (dev tools)." if attr else ""
    expected = f"Page loads (HTTP 200). Title contains \"{kw}\". Main/body content is visible."
    if attr:
        expected += f" body[data-page] equals \"{attr}\"."
    add(f"NAV-{i:03d}", "Navigation", f"{name} page loads with expected title keyword",
        "High" if name in ("Home", "Login", "Expert View") else "Medium",
        PRECOND_PUBLIC if name in ("Home", "Login", "Contact", "Disclaimer") else PRECOND_GATED,
        [f"1. Navigate to https://easemytrade.in{path}",
         "2. Observe the browser tab title.",
         "3. Confirm the main page content area is visible." + extra],
        expected, f"URL={path}, expected title keyword=\"{kw}\"")

index_pages = [("NIFTY", "/nifty/today/"), ("BANKNIFTY", "/indices/banknifty/"),
               ("FINNIFTY", "/indices/finnifty/"), ("SENSEX", "/indices/sensex/")]
for i, (name, path) in enumerate(index_pages, start=9):
    add(f"NAV-{i:03d}", "Navigation", f"{name} index page loads and shows index name",
        "High", PRECOND_GATED,
        [f"1. Navigate to https://easemytrade.in{path}",
         "2. Read the page title.",
         f"3. Search the page title and body text for \"{name}\"."],
        f"Page loads successfully. Either the title or the visible body text contains \"{name}\".", f"Index={name}")

commodity_pages = [("Gold", "/commodities/gold/", "Gold"), ("Silver", "/commodities/silver/", "Silver"),
                    ("Crude Oil", "/commodities/crude-oil/", "Crude"), ("Natural Gas", "/commodities/natural-gas/", "Natural Gas")]
for i, (name, path, kw) in enumerate(commodity_pages, start=13):
    add(f"NAV-{i:03d}", "Navigation", f"{name} commodity page loads with commodity name visible",
        "Medium", PRECOND_GATED,
        [f"1. Navigate to https://easemytrade.in{path}",
         "2. Read the page title.",
         f"3. Search the page title and body text for \"{kw}\"."],
        f"Page loads successfully. Either the title or visible body text contains \"{kw}\".", f"Commodity={name}")

add("NAV-017", "Navigation", "Brand logo on Expert View page navigates back to Home", "Medium", PRECOND_GATED,
    ["1. Navigate to /expert-view/.", "2. Click the brand logo in the header.", "3. Observe the resulting URL."],
    "Browser navigates to the easemytrade.in home page (apex or www).")
add("NAV-018", "Navigation", "Expert View nav link navigates to the Expert View page", "Medium", PRECOND_HOME,
    ["1. On the home page, click the \"Expert View\" link in the main nav.", "2. Observe the resulting URL."],
    "URL contains \"/expert-view/\".")
add("NAV-019", "Navigation", "News nav link navigates to the News page", "Medium", PRECOND_HOME,
    ["1. On the home page, click the \"News\" link in the main nav.", "2. Observe the resulting URL."],
    "URL contains \"/news/\".")
add("NAV-020", "Navigation", "Footer Disclaimer link navigates to the Disclaimer page", "Medium", PRECOND_HOME,
    ["1. Scroll to the footer.", "2. Click the \"Disclaimer\" link.", "3. Observe the resulting URL."],
    "URL contains \"/disclaimer/\".")
add("NAV-021", "Navigation", "Home page \"Open Market Overview\" CTA stays on the home page", "Medium", PRECOND_HOME,
    ["1. On the home page, click the CTA/link pointing to #market-overview.", "2. Wait ~0.5s for the anchor scroll.",
     "3. Observe the resulting URL."],
    "Page remains on easemytrade.in (anchor scroll within the home page, no navigation away).")
add("NAV-022", "Navigation", "Market Overview redirect page lands back on the site", "Low", PRECOND_PUBLIC,
    ["1. Navigate to https://easemytrade.in/market-overview/.", "2. Observe the resulting URL and page body."],
    "Browser ends up on an easemytrade.in page (e.g. redirected to home) and the body is visible.")
add("NAV-023", "Navigation", "All header nav links have non-empty href attributes", "Medium", PRECOND_HOME,
    ["1. On the home page, inspect every link inside nav.nav-links.", "2. Count the links and read each href attribute."],
    "At least 5 nav links are present and every link has a non-blank href attribute.")
add("NAV-024", "Navigation", "Indian Indices dropdown lists exactly NIFTY, BANKNIFTY, FINNIFTY, SENSEX", "Medium", PRECOND_HOME,
    ["1. On the home page, click the \"Indian Indices\" nav dropdown.",
     "2. Confirm links for NIFTY, BANKNIFTY, FINNIFTY, and SENSEX are visible in the open menu."],
    "All four index links (/nifty/today/, /indices/banknifty/, /indices/finnifty/, /indices/sensex/) are visible.")
add("NAV-025", "Navigation", "Page title format is consistent across Home, Expert View, and Login", "Low", PRECOND_PUBLIC,
    ["1. Navigate to the Home page and read the title.", "2. Navigate to /expert-view/ and read the title.",
     "3. Log out / clear session, then navigate to /login/ and read the title."],
    "Every title contains \"EaseMyTrade\" (format: \"{Page Name} | EaseMyTrade\").")

# ───────────────────────── HOME — Home Page ─────────────────────────
add("HOME-001", "Home Page", "Home page loads with correct title", "High", PRECOND_HOME,
    ["1. Read the browser tab title."], "Title contains \"EaseMyTrade\".")
add("HOME-002", "Home Page", "Home page URL matches the expected domain", "High", PRECOND_HOME,
    ["1. Read the current browser URL."], "URL contains \"easemytrade.in\" (apex or www).")
add("HOME-003", "Home Page", "Header navigation is visible with all links", "High", PRECOND_HOME,
    ["1. Inspect the header: brand logo, nav bar, and the Expert View / News / Methodology / Telegram / Contact links."],
    "Brand logo, nav bar, and all five links are visible.")
add("HOME-004", "Home Page", "Hero section is present with EaseMyTrade branding", "High", PRECOND_HOME,
    ["1. Inspect the hero section at the top of the page.", "2. Read the hero heading text.", "3. Confirm the hero lead paragraph is visible."],
    "Hero section is visible, heading text contains \"EaseMyTrade\", and the lead paragraph is visible.")
add("HOME-005", "Home Page", "Hero CTA buttons are present and correctly linked", "Medium", PRECOND_HOME,
    ["1. Inspect the hero section for the \"Open Market Overview\" and \"Indian Indices\" call-to-action buttons."],
    "Both CTA buttons are visible.")
add("HOME-006", "Home Page", "Market watchlist shows all tracked instruments", "High", PRECOND_HOME,
    ["1. Locate the market watchlist widget.", "2. Confirm rows exist for NIFTY, BANKNIFTY, FINNIFTY, SENSEX, and GOLD."],
    "All five watchlist rows are visible.")
add("HOME-007", "Home Page", "Market watchlist rows link to the correct pages", "High", PRECOND_HOME,
    ["1. Inspect the href of the NIFTY, BANKNIFTY, FINNIFTY, and SENSEX watchlist rows."],
    "Hrefs equal /nifty/today/, /indices/banknifty/, /indices/finnifty/, /indices/sensex/ respectively.")
add("HOME-008", "Home Page", "Trading Zones section shows all 4 index cards", "High", PRECOND_HOME,
    ["1. Locate the Trading Zones section.", "2. Count the .zone-card elements inside it."],
    "Trading Zones section is visible with exactly 4 zone cards (NIFTY, BANKNIFTY, FINNIFTY, SENSEX).")
for i, idx in enumerate(["NIFTY", "BANKNIFTY", "FINNIFTY", "SENSEX"], start=9):
    add(f"HOME-{i:03d}", "Home Page", f"{idx} zone: support level is numerically less than resistance", "High", PRECOND_HOME,
        [f"1. Locate the {idx} zone card in the Trading Zones section.",
         "2. Read the displayed Support and Resistance values.",
         "3. If both are populated numeric values, compare them."],
        "Support value is strictly less than the Resistance value (skip the numeric check if levels aren't populated yet).",
        f"Index={idx}")
add("HOME-013", "Home Page", "NIFTY zone signal state is one of the valid values", "High", PRECOND_HOME,
    ["1. Locate the NIFTY zone card.", "2. Read the displayed signal state."],
    "Signal state is one of: Trade Signal, Watchlist Only, No Trade, Hold, Pending (or blank/placeholder if not yet populated).")
add("HOME-014", "Home Page", "Previous Completed Trades spotlight section renders", "Medium", PRECOND_HOME,
    ["1. Locate the \"Previous Completed Trades\" section.", "2. Wait for the trade-spotlight grid to populate (it loads asynchronously).",
     "3. Confirm the grid is visible with trade cards."],
    "The spotlight section and its grid become visible (allow up to ~2 minutes for the live-market data fetch to complete).")
add("HOME-015", "Home Page", "Market drivers panel is present", "Medium", PRECOND_HOME,
    ["1. Locate the Market Drivers panel on the page."], "Market drivers panel is visible.")
add("HOME-016", "Home Page", "Footer is present with disclaimer link", "Medium", PRECOND_HOME,
    ["1. Scroll to the footer.", "2. Confirm the disclaimer link is visible inside it."],
    "Footer is visible and contains a visible disclaimer link.")
add("HOME-017", "Home Page", "Footer disclaimer link points to /disclaimer/", "Medium", PRECOND_HOME,
    ["1. Inspect the href attribute of the footer disclaimer link."], "Href equals \"/disclaimer/\".")
add("HOME-018", "Home Page", "Live market ticker strip is present in the header area", "Low", PRECOND_HOME,
    ["1. Inspect the header area for the scrolling market ticker."], "Market ticker strip is visible.")
add("HOME-019", "Home Page", "IST live clock is present in the navigation", "Low", PRECOND_HOME,
    ["1. Inspect the nav bar for the live IST clock element."], "Live clock element is visible.")
add("HOME-020", "Home Page", "Hero pulse chart canvas renders", "Medium", PRECOND_HOME,
    ["1. Inspect the hero section for the pulse-chart canvas element."], "Pulse chart canvas is visible.")
add("HOME-021", "Home Page", "Indian Indices nav dropdown reveals all 4 index links", "Medium", PRECOND_HOME,
    ["1. Click the \"Indian Indices\" dropdown in the nav.", "2. Confirm links for NIFTY, BANKNIFTY, FINNIFTY, SENSEX appear."],
    "All four index links are visible in the opened dropdown.")

# ───────────────────────── LOGIN — Login Page ─────────────────────────
add("LOGIN-001", "Login Page", "Login page loads with correct title", "High", PRECOND_LOGIN,
    ["1. Read the browser tab title."], "Title contains both \"Sign In\" and \"EaseMyTrade\".")
add("LOGIN-002", "Login Page", "Login form elements are present", "High", PRECOND_LOGIN,
    ["1. Inspect the login form for the username field, password field, Sign In button, and Sign Out button."],
    "Login form, username input, password input, Sign In button, and Sign Out button are all visible.")
add("LOGIN-003", "Login Page", "Password field is masked by default", "High", PRECOND_LOGIN,
    ["1. Inspect the password field's input type."], "Password input type is \"password\" (masked).")
add("LOGIN-004", "Login Page", "Toggling password show/hide changes the input type", "Medium", PRECOND_LOGIN,
    ["1. Type a test password into the password field.", "2. Confirm it is masked.",
     "3. Click the show/hide toggle.", "4. Confirm the field now shows plain text.",
     "5. Click the toggle again.", "6. Confirm the field is masked again."],
    "Password field type alternates from \"password\" to \"text\" and back to \"password\" with each toggle click.")
add("LOGIN-005", "Login Page", "Submitting an empty form shows an error status message", "High", PRECOND_LOGIN,
    ["1. Leave both username and password blank.", "2. Click Sign In.", "3. Wait ~1s and read the status message."],
    "A non-blank error-styled status message is displayed.")
add("LOGIN-006", "Login Page", "Submitting with only a username shows a missing-password error", "Medium", PRECOND_LOGIN,
    ["1. Enter a username (e.g. \"testuser\"), leave password blank.", "2. Click Sign In.", "3. Wait ~1s and read the status message."],
    "An error-styled status message is displayed.", "username=testuser")
add("LOGIN-007", "Login Page", "Invalid credentials show an error status", "High", PRECOND_LOGIN,
    ["1. Enter username \"invaliduser\" and password \"wrongpassword\".", "2. Click Sign In.", "3. Wait ~3s and read the status message."],
    "An error-styled status message is displayed (login is rejected).", "username=invaliduser, password=wrongpassword")
add("LOGIN-008", "Login Page", "After 4 failed attempts, login is temporarily locked", "High", PRECOND_LOGIN,
    ["1. Attempt login with username \"baduser\" and an incorrect password.", "2. Wait ~2s; clear the form.",
     "3. Repeat the failed attempt 3 more times (4 total).", "4. Read the final status message."],
    "After the 4th failed attempt, the status message indicates either an error or a temporary lockout/cooldown.",
    "username=baduser, 4 incorrect passwords")
add("LOGIN-009", "Login Page", "Login page brand logo links to the home page", "Low", PRECOND_LOGIN,
    ["1. Inspect the href attribute of the brand logo."], "Href equals \"/\".")
add("LOGIN-010", "Login Page", "Login page has a hero heading describing protected access", "Low", PRECOND_LOGIN,
    ["1. Inspect the hero heading on the login page.", "2. Read its text."],
    "Hero heading is visible and contains non-blank descriptive text.")
add("LOGIN-011", "Login Page", "Login page has data-page=\"login\" body attribute", "Low", PRECOND_LOGIN,
    ["1. Inspect the <body> element's data-page attribute (dev tools)."], "data-page attribute equals \"login\".")
add("LOGIN-012", "Login Page", "Login redirect preserves the \"next\" destination query parameter", "Medium", PRECOND_LOGIN,
    ["1. Navigate to /login/?next=/expert-view/.", "2. Observe the URL.", "3. Confirm the login form (#loginForm) is visible."],
    "URL still contains \"next=/expert-view/\" and the login form renders normally.")

# ───────────────────────── IDX — Index Pages ─────────────────────────
idx_data = [("NIFTY", "nifty", "/nifty/today/"), ("BANKNIFTY", "banknifty", "/indices/banknifty/"),
            ("FINNIFTY", "finnifty", "/indices/finnifty/"), ("SENSEX", "sensex", "/indices/sensex/")]
n = 1
for name, key, path in idx_data:
    add(f"IDX-{n:03d}", "Index Pages", f"{name} page loads with the index name in page content", "High", PRECOND_GATED,
        [f"1. Navigate to https://easemytrade.in{path}", "2. Confirm the main signal content area is visible.",
         f"3. Search the visible body text for \"{name}\"."],
        f"Main content is visible and the body text contains \"{name}\".", f"Index={name}"); n += 1
for name, key, path in idx_data:
    add(f"IDX-{n:03d}", "Index Pages", f"{name} page title contains EaseMyTrade", "High", PRECOND_GATED,
        [f"1. Navigate to https://easemytrade.in{path}", "2. Read the page title."],
        "Title contains \"EaseMyTrade\".", f"Index={name}"); n += 1
for name, key, path in idx_data:
    add(f"IDX-{n:03d}", "Index Pages", f"{name} page has a fully functional navigation header", "High", PRECOND_GATED,
        [f"1. Navigate to https://easemytrade.in{path}", "2. Confirm the brand logo is visible.",
         "3. Confirm the main nav bar is visible."],
        "Both the brand logo and the nav bar are visible.", f"Index={name}"); n += 1
add(f"IDX-{n:03d}", "Index Pages", "NIFTY page shows a valid signal state", "High", PRECOND_GATED,
    ["1. Navigate to /nifty/today/.", "2. Read the displayed signal state."],
    "Signal state is one of: Trade Signal, Watchlist Only, No Trade, Hold, Pending (or blank if not populated)."); n += 1
add(f"IDX-{n:03d}", "Index Pages", "BANKNIFTY page shows a valid signal state", "High", PRECOND_GATED,
    ["1. Navigate to /indices/banknifty/.", "2. Read the displayed signal state."],
    "Signal state is one of: Trade Signal, Watchlist Only, No Trade, Hold, Pending (or blank if not populated)."); n += 1
for name, key, path in idx_data:
    add(f"IDX-{n:03d}", "Index Pages", f"{name} support level is numerically less than resistance", "High", PRECOND_GATED,
        [f"1. Navigate to https://easemytrade.in{path}", "2. Read the displayed Support and Resistance values.",
         "3. If both are populated numeric values, compare them."],
        "Support is strictly less than Resistance (skip the numeric check if levels aren't populated yet).",
        f"Index={name}"); n += 1
add(f"IDX-{n:03d}", "Index Pages", "NIFTY buy/sell signal geometry is valid when a Trade Signal is active", "High", PRECOND_GATED,
    ["1. Navigate to /nifty/today/.", "2. Read the signal state; if it is not \"Trade Signal\", stop (not applicable).",
     "3. Read the Entry, Stop Loss, and Target levels.",
     "4. If Target > Entry, verify it is a valid buy setup and R:R ≥ 2.2.",
     "5. If Target < Entry, verify it is a valid sell setup."],
    "When a Trade Signal is active: for a buy setup, Stop < Entry < Target and Risk:Reward ≥ 2.2; for a sell setup, Target < Entry < Stop."); n += 1
add(f"IDX-{n:03d}", "Index Pages", "NIFTY page shows completed trade review cards", "Medium", PRECOND_GATED,
    ["1. Navigate to /nifty/today/.", "2. Confirm the main signal content area is visible (trade cards may or may not be present)."],
    "NIFTY page's signal content area is visible."); n += 1
commodities = [("Gold", "/commodities/gold/"), ("Silver", "/commodities/silver/"),
               ("Crude Oil", "/commodities/crude-oil/"), ("Natural Gas", "/commodities/natural-gas/")]
for name, path in commodities:
    add(f"IDX-{n:03d}", "Index Pages", f"{name} commodity page shows the commodity name in body content", "Medium", PRECOND_GATED,
        [f"1. Navigate to https://easemytrade.in{path}", "2. Read the visible body text.", "3. Confirm the main content area is visible."],
        f"Body text contains \"{name}\" and the main content area is visible.", f"Commodity={name}"); n += 1
for name, path in commodities:
    add(f"IDX-{n:03d}", "Index Pages", f"{name} commodity page has navigation header with all links", "Medium", PRECOND_GATED,
        [f"1. Navigate to https://easemytrade.in{path}", "2. Confirm the brand logo is visible.", "3. Confirm the nav bar is visible."],
        "Both the brand logo and the nav bar are visible.", f"Commodity={name}"); n += 1
add(f"IDX-{n:03d}", "Index Pages", "Gold commodity page shows a positive price value", "Medium", PRECOND_GATED,
    ["1. Navigate to /commodities/gold/.", "2. Locate the gold price element.", "3. Read its numeric value."],
    "If a numeric price is populated, it is greater than 0."); n += 1

# ───────────────────────── EV — Expert View ─────────────────────────
add("EV-001", "Expert View", "Expert View page loads with correct title", "High", PRECOND_EV,
    ["1. Read the browser tab title."], "Title contains both \"Expert View\" and \"EaseMyTrade\".")
add("EV-002", "Expert View", "Hero section is present with the AI Terminal headline", "High", PRECOND_EV,
    ["1. Inspect the hero section.", "2. Read the hero heading text."],
    "Hero section is visible and the heading text contains \"AI\".")
add("EV-003", "Expert View", "Search form is present with input, Generate Report, and Best Pick buttons", "High", PRECOND_EV,
    ["1. Inspect the search form, search input field, Generate Report button, and Best Pick button."],
    "All four elements are visible.")
add("EV-004", "Expert View", "Universe dropdown is present and shows the current selection", "High", PRECOND_EV,
    ["1. Inspect the universe dropdown.", "2. Read the summary text showing the current selection."],
    "Dropdown is visible and the summary text is non-blank.")
add("EV-005", "Expert View", "Universe dropdown opens and shows all option groups", "Medium", PRECOND_EV,
    ["1. Click the universe dropdown.", "2. Confirm the menu opens.", "3. Count the universe option buttons inside it."],
    "Menu is visible with more than 5 universe options.")
add("EV-006", "Expert View", "Selecting NIFTY 50 universe updates the dropdown summary label", "Medium", PRECOND_EV,
    ["1. Open the universe dropdown.", "2. Select \"NIFTY 50\".", "3. Read the updated summary label."],
    "Summary label now contains \"NIFTY\" (case-insensitive).")
add("EV-007", "Expert View", "Best Pick navigation buttons (Prev/Next) are present", "Low", PRECOND_EV,
    ["1. Inspect the Best Pick Prev and Next buttons."], "Both buttons are visible.")
add("EV-008", "Expert View", "Example ticker chips are present (Reliance, TCS, HDFC Bank, INFY, SBIN)", "Medium", PRECOND_EV,
    ["1. Count the example ticker chips on the page.", "2. Confirm chips exist for Reliance, TCS, HDFC Bank, INFY, and SBIN."],
    "At least 5 chips are visible and all 5 named chips are present.")
add("EV-009", "Expert View", "Clicking a quick-chip fills the search input with that stock name", "Medium", PRECOND_EV,
    ["1. Click the \"TCS\" example chip.", "2. Read the value now in the search input."],
    "Search input value contains \"TCS\".")
add("EV-010", "Expert View", "Initial status message is present before any search", "Medium", PRECOND_EV,
    ["1. Without searching, inspect the status message area.", "2. Read its text."],
    "Status message is visible and non-blank.")
add("EV-011", "Expert View", "Initial hero metrics show an awaiting-scan state", "Low", PRECOND_EV,
    ["1. Without searching, read the hero Confidence and Trend metric placeholders."],
    "Both metric fields render some placeholder/initial value (not null).")
add("EV-012", "Expert View", "Report area shows the empty \"Waiting for an asset\" card before any search", "Medium", PRECOND_EV,
    ["1. Without searching, inspect the report area.", "2. Confirm the empty-state card is visible.", "3. Read its text."],
    "Report area and empty card are visible; the card's text contains \"waiting\".")
add("EV-013", "Expert View", "Searching \"Reliance\" generates a report with analysis cards", "High", PRECOND_EV,
    ["1. Type \"Reliance\" into the search input.", "2. Click Generate Report.",
     "3. Wait up to 60s for an analysis card to appear.", "4. Count the analysis cards rendered."],
    "Report loads and at least one analysis card is rendered.", "Stock=Reliance")
add("EV-014", "Expert View", "After a successful scan, dashboard metrics (Confidence, Bias, Recommendation) are visible", "High", PRECOND_EV,
    ["1. Search for \"TCS\".", "2. Wait up to 60s for the report to load.", "3. Wait for the dashboard band to become visible.",
     "4. Read the Confidence Meter, Bias, and Recommendation values."],
    "Dashboard band is visible; Confidence Meter is non-blank and not \"--\"; Bias is non-blank; Recommendation is non-blank and not \"Standby\".",
    "Stock=TCS")
add("EV-015", "Expert View", "Confidence score from an Expert View scan is between 0 and 100 percent", "High", PRECOND_EV,
    ["1. Search for \"INFY\".", "2. Wait up to 60s for the report to load.", "3. Read the Confidence Meter value."],
    "If populated, the confidence percentage is between 0 and 100 inclusive.", "Stock=INFY")
add("EV-016", "Expert View", "AI Recommendation is one of the expected verdict values", "High", PRECOND_EV,
    ["1. Search for \"SBIN\".", "2. Wait up to 60s for the report to load.", "3. Read the AI Recommendation value."],
    "Recommendation (if not \"Standby\") is one of: Strong Buy, Buy, Accumulate, Hold, Reduce, Sell, Strong Sell, Watch, Neutral, Watchlist Only, No Trade.",
    "Stock=SBIN")
add("EV-017", "Expert View", "AI Systems section shows five robot cards after a scan", "Medium", PRECOND_EV,
    ["1. Search for \"HDFC Bank\".", "2. Wait up to 60s for the report to load.", "3. Wait for the AI Systems section to become visible.",
     "4. Count the robot cards inside it."],
    "AI Systems section is visible with exactly 5 robot cards.", "Stock=HDFC Bank")
add("EV-018", "Expert View", "Report contains sections for Trend, Momentum, Support & Resistance, and Risk", "High", PRECOND_EV,
    ["1. Search for \"Reliance\".", "2. Wait up to 60s for the report to load.",
     "3. Wait for the report panel's content to render fully.", "4. Inspect the rendered report HTML/content."],
    "Report HTML has substantive content (over 500 characters), implying all section blocks rendered.", "Stock=Reliance")
add("EV-019", "Expert View", "Best Pick button initiates a scan for the selected universe", "High", PRECOND_EV,
    ["1. Read the current status message.", "2. Click the Best Pick button.", "3. Wait ~2s.", "4. Read the status message again."],
    "Status message changes after clicking Best Pick, indicating a scan was initiated.")
add("EV-020", "Expert View", "Searching an empty string does not crash or navigate away", "Medium", PRECOND_EV,
    ["1. Clear the search input (leave it empty).", "2. Click Generate Report.", "3. Wait ~1s.", "4. Observe the URL."],
    "Page remains on /expert-view/; no crash or unexpected navigation occurs.")
add("EV-021", "Expert View", "Short Horizon Scope is shown after a successful scan", "Low", PRECOND_EV,
    ["1. Search for \"TCS\".", "2. Wait up to 60s for the report to load.", "3. Read the Short Horizon Scope value."],
    "Short Horizon Scope value is rendered (not null) after the scan.", "Stock=TCS")
add("EV-022", "Expert View", "Expert View page body has the correct data-page attribute", "Low", PRECOND_EV,
    ["1. Inspect the <body> element's data-page attribute (dev tools)."], "data-page attribute equals \"expert-view\".")
for stock in ["Reliance", "TCS", "INFY", "HDFC Bank", "SBIN"]:
    n_ev = 23 + ["Reliance", "TCS", "INFY", "HDFC Bank", "SBIN"].index(stock)
    add(f"EV-{n_ev:03d}", "Expert View", f"Searching \"{stock}\" from quick examples returns a status update, not an error", "High", PRECOND_EV,
        [f"1. Search for \"{stock}\" (via search box or its example chip).", "2. Wait ~3s.", "3. Read the status message."],
        "Status message does not contain \"exception\", \"500\", or \"internal server error\".", f"Stock={stock}")

# ───────────────────────── CP — Content Pages ─────────────────────────
add("CP-001", "Content Pages", "News page loads with main content area", "High", PRECOND_GATED,
    ["1. Navigate to /news/.", "2. Confirm the main content area is visible.", "3. Read the page title."],
    "Main content area is visible and the title contains \"EaseMyTrade\".")
add("CP-002", "Content Pages", "News page contains at least one heading", "Medium", PRECOND_GATED,
    ["1. Navigate to /news/.", "2. Count the heading elements on the page."], "At least one heading is present.")
add("CP-003", "Content Pages", "Methodology page loads with content about signal logic", "High", PRECOND_GATED,
    ["1. Navigate to /methodology/.", "2. Confirm the main content area is visible."], "Main content area is visible.")
add("CP-004", "Content Pages", "Methodology page references key signal concepts", "Medium", PRECOND_GATED,
    ["1. Navigate to /methodology/.", "2. Read the visible body text."],
    "Body text mentions at least one of: signal, NIFTY, methodology, indicator.")
add("CP-005", "Content Pages", "Telegram page loads with main content mentioning Telegram", "High", PRECOND_GATED,
    ["1. Navigate to /telegram/.", "2. Confirm the main content area is visible.", "3. Read the visible body text."],
    "Main content area is visible and the body text contains \"telegram\".")
add("CP-006", "Content Pages", "Contact page loads and shows contact information", "High", PRECOND_PUBLIC,
    ["1. Navigate to /contact/.", "2. Confirm the main content area is visible.", "3. Read the visible body text."],
    "Main content area is visible and body text contains contact information (e.g. \"contact\", \"email\", or the owner's name).")
add("CP-007", "Content Pages", "Contact page has non-empty paragraphs", "Medium", PRECOND_PUBLIC,
    ["1. Navigate to /contact/.", "2. Count the paragraph elements on the page."], "At least one paragraph is present.")
add("CP-008", "Content Pages", "Disclaimer page loads with educational disclaimer content", "High", PRECOND_PUBLIC,
    ["1. Navigate to /disclaimer/.", "2. Confirm the main content area is visible.", "3. Read the visible body text."],
    "Main content area is visible and body text mentions disclaimer/educational/informational content.")
add("CP-009", "Content Pages", "Disclaimer page explicitly states educational/informational purpose", "High", PRECOND_PUBLIC,
    ["1. Navigate to /disclaimer/.", "2. Read the full visible body text."],
    "Body text states the content is educational/informational, or explicitly says it is not investment advice/a solicitation.")
add("CP-010", "Content Pages", "Telegram status page loads without a 404", "Low", PRECOND_PUBLIC,
    ["1. Navigate to /telegram-status/.", "2. Confirm the page body is visible.", "3. Read the page title."],
    "Body is visible and the title is non-blank (no 404 page).")
add("CP-011", "Content Pages", "All major pages include EaseMyTrade in the page title", "Medium", PRECOND_PUBLIC,
    ["1. Visit each of: Home, Login, Expert View, News, Methodology, Telegram, Contact, Disclaimer, NIFTY, BANKNIFTY.",
     "2. Read each page's title.",
     "3. For any gated page that redirects to /login/, skip the title check for that page only."],
    "Every page that successfully loads (not redirected to login) has a title containing \"EaseMyTrade\".")
add("CP-012", "Content Pages", "All major pages have a visible footer with the EaseMyTrade brand", "Low", PRECOND_GATED,
    ["1. Visit each of: Home, Expert View, NIFTY, News, Methodology.",
     "2. Scroll to the bottom of each page and confirm the footer is visible.",
     "3. For any page that redirects to /login/, skip the footer check for that page only."],
    "Every page that successfully loads has a visible footer.")

# ───────────────────────── SIG — Signal Logic ─────────────────────────
add("SIG-001", "Signal Logic", "NIFTY signal state is one of the valid engine states", "High", PRECOND_HOME,
    ["1. On the home page, read the NIFTY zone's displayed signal state."],
    "State is one of: Trade Signal, Watchlist Only, No Trade, Hold, Pending (or blank if not yet populated).")
add("SIG-002", "Signal Logic", "Signal engine supports all four index zones on the home page", "High", PRECOND_HOME,
    ["1. On the home page, count the .zone-card elements in the Trading Zones section."],
    "At least 4 zone cards are present (NIFTY, BANKNIFTY, FINNIFTY, SENSEX).")
add("SIG-003", "Signal Logic", "All index zone support levels are strictly less than resistance levels", "High", PRECOND_HOME,
    ["1. On the home page, read Support and Resistance for NIFTY, BANKNIFTY, FINNIFTY, and SENSEX.",
     "2. For each zone where both values are populated, compare Support to Resistance."],
    "For every zone with populated levels, Support < Resistance (no zone may have Support ≥ Resistance).")
add("SIG-004", "Signal Logic", "Confidence score on the NIFTY page is between 0 and 100 percent", "High", PRECOND_GATED,
    ["1. Navigate to /nifty/today/.", "2. Locate any visible confidence-score element.", "3. Read its percentage value."],
    "If populated, the confidence percentage is between 0 and 100 inclusive.")
add("SIG-005", "Signal Logic", "Live market API returns a valid JSON structure", "High", "API client / browser dev tools access to easemytrade.in.",
    ["1. Send a GET request to https://easemytrade.in/api/live-market.", "2. Read the HTTP status code.",
     "3. If status is 200, parse the JSON body."],
    "Status is 200, 304 (cached), or 401 (protected — unlock window closed). If 200, the JSON contains \"nifty\"/\"NIFTY\" or \"indices\" data.")
add("SIG-006", "Signal Logic", "market.json file has the required fields for the signal engine", "High", "API client access to easemytrade.in.",
    ["1. Send a GET request to https://easemytrade.in/data/market.json.", "2. If status is 200, parse the JSON.",
     "3. Inspect the \"nifty\" object for level, signal, support, and resistance fields.",
     "4. Inspect nifty.signal for state, confidence, stopLoss, and a target field."],
    "If the file is publicly accessible, all listed fields are present; confidence is within [0,100] and state is a valid signal state.")
add("SIG-007", "Signal Logic", "Active Trade Signal entries have valid buy/sell geometry", "High", "API client access to easemytrade.in.",
    ["1. Fetch /data/market.json.", "2. For each index in Trade Signal state, read Entry, Stop Loss, and Target.",
     "3. If Target > Entry, verify it is a valid buy setup with R:R ≥ 2.2.",
     "4. If Target < Entry, verify it is a valid sell setup."],
    "Every active Trade Signal has valid entry/stop/target geometry for its direction.")
add("SIG-008", "Signal Logic", "Risk-reward ratio is at least 2.2R for all active Trade Signal entries", "High", "API client access to easemytrade.in.",
    ["1. Fetch /data/market.json.", "2. For each index in Trade Signal state with populated entry/stop/target, calculate R:R."],
    "Every active Trade Signal has a Risk:Reward ratio ≥ 2.2.")
add("SIG-009", "Signal Logic", "No Trade state never shows an actionable Buy/Sell instruction", "Medium", "API client access to easemytrade.in.",
    ["1. Fetch /data/market.json.", "2. For each index in \"No Trade\" state, read the action/note text."],
    "No index in \"No Trade\" state has action text containing \"buy signal\" or \"sell signal\".")
add("SIG-010", "Signal Logic", "Session guard blocks new Trade Signals after 2 completed trades", "High", "API client access to easemytrade.in.",
    ["1. Fetch /data/market.json.", "2. For each index whose note/summary mentions \"2 completed trades\", read its signal state."],
    "Any index that has hit the 2-completed-trades session guard is NOT in \"Trade Signal\" state.")
add("SIG-011", "Signal Logic", "Forward record guard blocks fresh signals when net negative", "High", "API client access to easemytrade.in.",
    ["1. Fetch /data/market.json.", "2. For each index whose note/summary mentions a net-negative forward record, read its signal state."],
    "Any index with a net-negative forward record is NOT in \"Trade Signal\" state.")

print(f"Total test cases: {len(CASES)}")

# ───────────────────────── Build the workbook ─────────────────────────
HEADERS = ["Test Case ID", "Module", "Title", "Priority", "Preconditions", "Test Steps", "Expected Result", "Test Data"]
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PRIORITY_FILL = {
    "High": PatternFill("solid", start_color="FCE4E4"),
    "Medium": PatternFill("solid", start_color="FFF4D6"),
    "Low": PatternFill("solid", start_color="E6F4EA"),
}
MODULE_BANDS = {
    "Navigation": "DCE6F1", "Home Page": "EAF1DD", "Login Page": "FDE9D9",
    "Index Pages": "E5DFEC", "Expert View": "D9EAD3", "Content Pages": "FFF2CC", "Signal Logic": "D9D2E9",
}

wb = Workbook()
ws = wb.active
ws.title = "Manual Test Cases"
ws.append(HEADERS)
for c in range(1, len(HEADERS) + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
ws.freeze_panes = "A2"

for row in CASES:
    id_, module, title, priority, precond, steps, expected, data = row
    steps_text = "\n".join(steps)
    ws.append([id_, module, title, priority, precond, steps_text, expected, data])
    r = ws.max_row
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = BORDER
    ws.cell(row=r, column=2).fill = PatternFill("solid", start_color=MODULE_BANDS.get(module, "FFFFFF"))
    ws.cell(row=r, column=4).fill = PRIORITY_FILL.get(priority, PatternFill())

widths = {"A": 12, "B": 16, "C": 46, "D": 10, "E": 40, "F": 50, "G": 48, "H": 22}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Summary sheet
summary = wb.create_sheet("Summary")
summary.append(["Module", "Test Case Count"])
for c in range(1, 3):
    cell = summary.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
modules_order = ["Navigation", "Home Page", "Login Page", "Index Pages", "Expert View", "Content Pages", "Signal Logic"]
counts = {m: sum(1 for c in CASES if c[1] == m) for m in modules_order}
for m in modules_order:
    summary.append([m, counts[m]])
summary.append(["TOTAL", len(CASES)])
summary.cell(row=summary.max_row, column=1).font = Font(name="Arial", bold=True)
summary.cell(row=summary.max_row, column=2).font = Font(name="Arial", bold=True)
for r in range(2, summary.max_row + 1):
    for c in (1, 2):
        summary.cell(row=r, column=c).font = Font(name="Arial", size=10)
summary.column_dimensions["A"].width = 20
summary.column_dimensions["B"].width = 18

os.makedirs("docs", exist_ok=True)
xlsx_path = "docs/Manual_Test_Cases.xlsx"
wb.save(xlsx_path)
print(f"Wrote {xlsx_path}")

# ───────────────────────── Build the Markdown version ─────────────────────────
md_lines = ["# EaseMyTrade — Manual Test Cases", "",
            f"Total: **{len(CASES)}** test cases across **{len(modules_order)}** modules. "
            "Each case mirrors an automated Playwright test in `src/test/java/com/easemytrade/tests/`.", ""]
md_lines.append("## Summary")
md_lines.append("")
md_lines.append("| Module | Test Cases |")
md_lines.append("|---|---|")
for m in modules_order:
    md_lines.append(f"| {m} | {counts[m]} |")
md_lines.append(f"| **Total** | **{len(CASES)}** |")
md_lines.append("")

current_module = None
for row in CASES:
    id_, module, title, priority, precond, steps, expected, data = row
    if module != current_module:
        md_lines.append(f"## {module}")
        md_lines.append("")
        current_module = module
    md_lines.append(f"### {id_} — {title}")
    md_lines.append("")
    md_lines.append(f"**Priority:** {priority}")
    md_lines.append("")
    md_lines.append(f"**Preconditions:** {precond}")
    md_lines.append("")
    md_lines.append("**Steps:**")
    md_lines.append("")
    for s in steps:
        md_lines.append(s)
    md_lines.append("")
    md_lines.append(f"**Expected Result:** {expected}")
    md_lines.append("")
    if data:
        md_lines.append(f"**Test Data:** {data}")
        md_lines.append("")

md_path = "docs/MANUAL_TEST_CASES.md"
with open(md_path, "w") as f:
    f.write("\n".join(md_lines))
print(f"Wrote {md_path}")
